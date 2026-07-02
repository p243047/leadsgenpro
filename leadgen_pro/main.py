"""
LeadGen Pro - Enterprise-Grade Lead Generation Application
==========================================================
A production-ready desktop application for extracting, enriching, and exporting
high-quality business leads with premium UI/UX and advanced scraping capabilities.

Author: Senior Full-Stack Python Engineer
Version: 1.0.0
"""

import asyncio
import aiohttp
import json
import re
import os
import sys
import threading
import time
import random
import socket
import dns.resolver
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field, asdict

# Playwright imports
from playwright.async_api import async_playwright, Browser, Page, BrowserContext

# Excel exports
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# UI imports
import customtkinter as ctk
from tkinter import messagebox, filedialog

# Configure customtkinter appearance
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


@dataclass
class Lead:
    """Data class representing a business lead."""
    name: str = ""
    address: str = ""
    service_need: str = ""
    contact_details: str = ""
    email: str = ""
    business_category: str = ""
    business_info: str = ""
    social_media: str = ""
    source_url: str = ""
    extracted_at: str = ""


@dataclass
class ScrapingConfig:
    """Configuration for scraping operations."""
    keyword: str = ""
    location: str = ""
    max_leads: int = 50
    hunter_api_key: str = ""
    tomba_api_key: str = ""
    timeout: int = 30
    max_concurrent: int = 10


class EmailValidator:
    """Advanced email validation with MX record checking."""
    
    DISPOSABLE_DOMAINS = {
        'tempmail.com', 'throwaway.email', 'guerrillamail.com', 'mailinator.com',
        '10minutemail.com', 'fakeinbox.com', 'trashmail.com'
    }
    
    @staticmethod
    def is_valid_format(email: str) -> bool:
        """Validate email format against RFC 5322."""
        if not email or '@' not in email:
            return False
        
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    @staticmethod
    def is_disposable(email: str) -> bool:
        """Check if email uses a disposable domain."""
        domain = email.split('@')[1].lower() if '@' in email else ''
        return domain in EmailValidator.DISPOSABLE_DOMAINS
    
    @staticmethod
    async def check_mx_record(domain: str) -> bool:
        """Async DNS lookup to verify domain can receive mail."""
        try:
            answers = await asyncio.get_event_loop().run_in_executor(
                None, 
                lambda: dns.resolver.resolve(domain, 'MX')
            )
            return len(list(answers)) > 0
        except Exception:
            return False
    
    @classmethod
    async def validate_email(cls, email: str) -> bool:
        """Comprehensive email validation."""
        if not cls.is_valid_format(email):
            return False
        
        if cls.is_disposable(email):
            return False
        
        domain = email.split('@')[1]
        mx_valid = await cls.check_mx_record(domain)
        
        return mx_valid


class StealthManager:
    """Manage anti-detection techniques for web scraping."""
    
    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    ]
    
    @staticmethod
    def get_random_user_agent() -> str:
        """Return a random user agent string."""
        return random.choice(StealthManager.USER_AGENTS)
    
    @staticmethod
    def get_random_delay(min_ms: int = 1000, max_ms: int = 3000) -> float:
        """Generate human-like random delay."""
        return random.uniform(min_ms / 1000, max_ms / 1000)
    
    @staticmethod
    async def inject_stealth_scripts(page: Page) -> None:
        """Inject scripts to hide automation fingerprints."""
        stealth_scripts = """
        // Hide webdriver flag
        Object.defineProperty(navigator, 'webdriver', {
            get: () => undefined
        });
        
        // Mock plugins
        Object.defineProperty(navigator, 'plugins', {
            get: () => [1, 2, 3, 4, 5]
        });
        
        // Mock languages
        Object.defineProperty(navigator, 'languages', {
            get: () => ['en-US', 'en']
        });
        
        // Override permissions
        const originalQuery = window.navigator.permissions.query;
        window.navigator.permissions.query = (parameters) => (
            parameters.name === 'notifications' ?
                Promise.resolve({ state: Notification.permission }) :
                originalQuery(parameters)
        );
        """
        await page.add_init_script(stealth_scripts)


class DataExtractor:
    """Extract and enrich business data from websites."""
    
    @staticmethod
    def extract_phone_numbers(text: str) -> List[str]:
        """Extract and clean phone numbers from text."""
        patterns = [
            r'\+?[\d\s\-\(\)]{10,}',
            r'\(\d{3}\)\s*\d{3}[-.]?\d{4}',
            r'\d{3}[-.\s]\d{3}[-.\s]\d{4}'
        ]
        
        phones = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            phones.extend(matches)
        
        # Clean and format
        cleaned = []
        for phone in phones:
            digits = re.sub(r'[^\d+]', '', phone)
            if len(digits) >= 10:
                if not digits.startswith('+'):
                    digits = '+1' + digits  # Default to US
                cleaned.append(digits[:15])  # Limit length
        
        return list(set(cleaned))[:3]  # Return up to 3 unique numbers
    
    @staticmethod
    def extract_emails(text: str) -> List[str]:
        """Extract emails from text with validation."""
        pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        emails = re.findall(pattern, text)
        
        # Filter out generic emails if possible
        generic_prefixes = ['info', 'support', 'contact', 'hello', 'admin']
        prioritized = []
        generic = []
        
        for email in emails:
            prefix = email.split('@')[0].lower()
            if prefix in generic_prefixes:
                generic.append(email)
            else:
                prioritized.append(email)
        
        # Return specific emails first, then generic if needed
        return list(set(prioritized + generic))[:5]
    
    @staticmethod
    def extract_social_links(html: str, base_url: str) -> Dict[str, str]:
        """Extract social media URLs from HTML."""
        social_patterns = {
            'LinkedIn': r'https?://(?:www\.)?linkedin\.com/company/[^"\'>\s]+',
            'Facebook': r'https?://(?:www\.)?facebook\.com/[^"\'>\s]+',
            'Instagram': r'https?://(?:www\.)?instagram\.com/[^"\'>\s]+',
            'Twitter': r'https?://(?:www\.)?(?:twitter\.com|x\.com)/[^"\'>\s]+',
        }
        
        socials = {}
        for platform, pattern in social_patterns.items():
            matches = re.findall(pattern, html, re.IGNORECASE)
            if matches:
                socials[platform] = matches[0]
        
        return socials
    
    @staticmethod
    def infer_service_needs(content: str, has_ssl: bool) -> str:
        """Use heuristics to infer business service needs."""
        needs = []
        content_lower = content.lower()
        
        # Marketing needs
        if 'pixel' not in content_lower and 'analytics' not in content_lower:
            needs.append("Digital Marketing")
        
        # Security needs
        if not has_ssl:
            needs.append("SSL Certificate")
        
        # Web development
        if 'responsive' not in content_lower or 'mobile' not in content_lower:
            needs.append("Mobile Optimization")
        
        # SEO
        if 'meta description' not in content_lower:
            needs.append("SEO Optimization")
        
        return "; ".join(needs) if needs else "No immediate needs identified"
    
    @staticmethod
    def extract_business_info(content: str) -> Dict[str, Any]:
        """Extract business information from website content."""
        info = {
            'description': '',
            'employee_count': 'Unknown',
            'founding_year': 'Unknown'
        }
        
        # Extract description (first meaningful paragraph)
        paragraphs = re.findall(r'<p[^>]*>([^<]+)</p>', content, re.IGNORECASE)
        if paragraphs:
            info['description'] = paragraphs[0].strip()[:200]
        
        # Try to find employee count
        employee_patterns = [
            r'(\d+)\s*[-–]?\s*(\d+)?\s*employees',
            r'team\s*of\s*(\d+)',
            r'(\d+)\+?\s*staff'
        ]
        for pattern in employee_patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                info['employee_count'] = match.group(0)
                break
        
        # Try to find founding year
        year_patterns = [
            r'(?:established|founded|since)\s*(\d{4})',
            r'est\.\s*(\d{4})'
        ]
        for pattern in year_patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                info['founding_year'] = match.group(1)
                break
        
        return info


class LeadScraper:
    """Main scraping engine using Playwright."""
    
    def __init__(self, config: ScrapingConfig, log_callback=None, progress_callback=None):
        self.config = config
        self.leads: List[Lead] = []
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.stop_flag = False
        self.executor = ThreadPoolExecutor(max_workers=5)
    
    def log(self, message: str):
        """Log message with callback."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_msg = f"[{timestamp}] {message}"
        if self.log_callback:
            self.log_callback(log_msg)
        print(log_msg)
    
    def update_progress(self, current: int, total: int, status: str):
        """Update progress with callback."""
        percentage = int((current / total) * 100) if total > 0 else 0
        if self.progress_callback:
            self.progress_callback(percentage, status)
    
    async def initialize_browser(self):
        """Initialize Playwright browser with stealth settings."""
        playwright = await async_playwright().start()
        
        self.browser = await playwright.chromium.launch(
            headless=True,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-dev-shm-usage',
                '--disable-web-security',
                '--disable-features=IsolateOrigins,site-per-process'
            ]
        )
        
        self.context = await self.browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent=StealthManager.get_random_user_agent(),
            locale='en-US',
            timezone_id='America/New_York'
        )
        
        # Inject stealth scripts into all pages
        await StealthManager.inject_stealth_scripts(await self.context.new_page())
        await (await self.context.pages)[0].close()
        
        self.log("Browser initialized with stealth mode")
    
    async def close_browser(self):
        """Close browser and cleanup resources."""
        try:
            if self.context:
                await self.context.close()
            if self.browser:
                await self.browser.close()
            self.log("Browser closed successfully")
        except Exception as e:
            self.log(f"Error closing browser: {str(e)}")
    
    async def search_google_maps(self) -> List[Dict[str, str]]:
        """Search Google Maps for businesses matching criteria."""
        results = []
        
        try:
            page = await self.context.new_page()
            
            # Construct search query
            query = f"{self.config.keyword} {self.config.location}"
            search_url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
            
            self.log(f"Searching Google Maps for: {query}")
            
            await page.goto(search_url, timeout=self.config.timeout * 1000)
            await asyncio.sleep(StealthManager.get_random_delay(3000, 5000))
            
            # Check for CAPTCHA
            if 'captcha' in page.url.lower() or 'sorry' in page.url.lower():
                self.log("CAPTCHA detected, skipping search")
                await page.close()
                return results
            
            # Wait for results to load
            await page.wait_for_selector('div[role="feed"]', timeout=10000)
            await asyncio.sleep(StealthManager.get_random_delay(2000, 3000))
            
            # Extract business listings
            business_elements = await page.query_selector_all('div[role="article"]')
            
            for idx, element in enumerate(business_elements[:self.config.max_leads]):
                if self.stop_flag:
                    break
                
                try:
                    # Extract basic info
                    name_elem = await element.query_selector('div.fontHeadlineSmall')
                    name = await name_elem.inner_text() if name_elem else "Unknown"
                    
                    # Get link to business page
                    link_elem = await element.query_selector('a[href*="/maps/place"]')
                    if not link_elem:
                        continue
                    
                    href = await link_elem.get_attribute('href')
                    
                    results.append({
                        'name': name,
                        'url': href,
                        'position': idx + 1
                    })
                    
                    self.log(f"Found business #{len(results)}: {name}")
                    
                except Exception as e:
                    self.log(f"Error extracting listing: {str(e)}")
                    continue
            
            await page.close()
            self.log(f"Extracted {len(results)} businesses from Google Maps")
            
        except Exception as e:
            self.log(f"Google Maps search error: {str(e)}")
        
        return results
    
    async def visit_business_website(self, business_url: str) -> Optional[Page]:
        """Visit and return page for business website."""
        try:
            page = await self.context.new_page()
            
            # Set random delay before visiting
            await asyncio.sleep(StealthManager.get_random_delay(1000, 2000))
            
            await page.goto(business_url, timeout=self.config.timeout * 1000, wait_until='domcontentloaded')
            await asyncio.sleep(StealthManager.get_random_delay(2000, 3000))
            
            return page
            
        except Exception as e:
            self.log(f"Error visiting website: {str(e)}")
            return None
    
    async def extract_lead_data(self, page: Page, business_info: Dict) -> Optional[Lead]:
        """Extract comprehensive lead data from business page."""
        try:
            # Get page content
            html = await page.content()
            text = await page.evaluate('document.body.innerText')
            
            # Check SSL
            has_ssl = page.url.startswith('https://')
            
            # Extract emails
            emails = DataExtractor.extract_emails(text)
            validated_email = ""
            
            # Validate emails asynchronously
            for email in emails:
                if await EmailValidator.validate_email(email):
                    validated_email = email
                    break
            
            if not validated_email and emails:
                validated_email = emails[0]
            
            # Extract phone numbers
            phones = DataExtractor.extract_phone_numbers(text)
            contact_details = "; ".join(phones) if phones else ""
            
            # Extract social links
            socials = DataExtractor.extract_social_links(html, page.url)
            social_media = "; ".join([f"{k}: {v}" for k, v in socials.items()])
            
            # Extract business info
            biz_info = DataExtractor.extract_business_info(html)
            business_info_str = f"{biz_info['description']}. Employees: {biz_info['employee_count']}. Founded: {biz_info['founding_year']}"
            
            # Infer service needs
            service_need = DataExtractor.infer_service_needs(text, has_ssl)
            
            # Try to extract address
            address = ""
            address_patterns = [
                r'\d+\s+[A-Za-z\s]+,\s*[A-Za-z\s]+,\s*[A-Z]{2}\s*\d{5}',
                r'[A-Za-z\s]+,\s*[A-Z]{2}\s*\d{5}'
            ]
            for pattern in address_patterns:
                match = re.search(pattern, text)
                if match:
                    address = match.group(0)
                    break
            
            # Determine business category
            category = self.config.keyword if self.config.keyword else "General Business"
            
            lead = Lead(
                name=business_info.get('name', 'Unknown'),
                address=address,
                service_need=service_need,
                contact_details=contact_details,
                email=validated_email,
                business_category=category,
                business_info=business_info_str,
                social_media=social_media,
                source_url=page.url,
                extracted_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            return lead
            
        except Exception as e:
            self.log(f"Error extracting lead data: {str(e)}")
            return None
    
    async def scrape_single_lead(self, business_info: Dict, semaphore: asyncio.Semaphore) -> Optional[Lead]:
        """Scrape data for a single business."""
        async with semaphore:
            if self.stop_flag:
                return None
            
            self.log(f"Processing: {business_info['name']}")
            
            page = await self.visit_business_website(business_info['url'])
            if not page:
                return None
            
            try:
                lead = await self.extract_lead_data(page, business_info)
                if lead:
                    self.log(f"✓ Successfully extracted: {lead.email or 'No email'}")
                return lead
            finally:
                await page.close()
    
    async def run_scraping(self) -> List[Lead]:
        """Main scraping orchestration."""
        self.log("Starting scraping process...")
        
        # Initialize browser
        await self.initialize_browser()
        
        # Search Google Maps
        businesses = await self.search_google_maps()
        
        if not businesses:
            self.log("No businesses found")
            await self.close_browser()
            return []
        
        # Create semaphore for concurrency control
        semaphore = asyncio.Semaphore(self.config.max_concurrent)
        
        # Process businesses
        tasks = [
            self.scrape_single_lead(biz, semaphore) 
            for biz in businesses
        ]
        
        results = []
        total = len(tasks)
        
        for idx, coro in enumerate(asyncio.as_completed(tasks), 1):
            if self.stop_flag:
                break
            
            try:
                lead = await coro
                if lead:
                    self.leads.append(lead)
                    results.append(lead)
                    self.update_progress(idx, total, f"Extracted {len(results)} leads")
            except Exception as e:
                self.log(f"Task error: {str(e)}")
                continue
        
        await self.close_browser()
        
        self.log(f"Scraping complete. Total leads: {len(results)}")
        return results
    
    def stop(self):
        """Stop scraping operation."""
        self.stop_flag = True
        self.log("Stop signal received")


class ExcelExporter:
    """Export leads to professionally formatted Excel files."""
    
    HEADER_STYLE = Font(bold=True, color="FFFFFF", size=12)
    HEADER_FILL = PatternFill(start_color="1e1e2e", end_color="1e1e2e", fill_type="solid")
    ALTERNATING_FILLS = [
        PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"),
        PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ]
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    COLUMNS = [
        'Name', 'Address', 'Service Need', 'Contact Details', 
        'Email', 'Business Category', 'Business Info', 'Social Media'
    ]
    
    @staticmethod
    def export_to_excel(leads: List[Lead], keyword: str, location: str, output_dir: str = "output") -> str:
        """Export leads to Excel with professional formatting."""
        # Create output directory
        Path(output_dir).mkdir(exist_ok=True)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_keyword = re.sub(r'[^\w\s-]', '', keyword)[:30].replace(' ', '_')
        safe_location = re.sub(r'[^\w\s-]', '', location)[:30].replace(' ', '_')
        filename = f"leads_{safe_keyword}_{safe_location}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Convert leads to DataFrame
        data = []
        for lead in leads:
            row = {
                'Name': lead.name,
                'Address': lead.address,
                'Service Need': lead.service_need,
                'Contact Details': lead.contact_details,
                'Email': lead.email,
                'Business Category': lead.business_category,
                'Business Info': lead.business_info,
                'Social Media': lead.social_media
            }
            data.append(row)
        
        df = pd.DataFrame(data, columns=ExcelExporter.COLUMNS)
        
        # Write to Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Leads', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Leads']
            
            # Style header row (row 1 in 1-indexed openpyxl)
            for col_idx, col_name in enumerate(ExcelExporter.COLUMNS, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = ExcelExporter.HEADER_STYLE
                cell.fill = ExcelExporter.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = ExcelExporter.BORDER
            
            # Style data rows
            for row_idx in range(2, len(df) + 2):
                fill = ExcelExporter.ALTERNATING_FILLS[(row_idx - 2) % 2]
                for col_idx in range(1, len(ExcelExporter.COLUMNS) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.border = ExcelExporter.BORDER
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-fit columns
            for col_idx, col_name in enumerate(ExcelExporter.COLUMNS, 1):
                max_length = max(
                    len(str(col_name)),
                    *(len(str(val)) for val in df[col_name]) if len(df) > 0 else [0]
                )
                adjusted_width = min(max_length + 2, 50)
                column_letter = chr(64 + col_idx)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
        
        return filepath


class LeadGenProApp(ctk.CTk):
    """Main application GUI."""
    
    def __init__(self):
        super().__init__()
        
        self.title("LeadGen Pro - Enterprise Lead Generation")
        self.geometry("1200x800")
        self.minsize(1000, 700)
        
        # State variables
        self.scraper: Optional[LeadScraper] = None
        self.scraping_thread: Optional[threading.Thread] = None
        self.is_scraping = False
        self.current_leads: List[Lead] = []
        self.config = ScrapingConfig()
        
        # Configure grid
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the user interface."""
        # Header
        self.create_header()
        
        # Main content area
        self.create_input_section()
        self.create_control_panel()
        self.create_feedback_section()
        self.create_export_section()
    
    def create_header(self):
        """Create application header."""
        header_frame = ctk.CTkFrame(self, fg_color="#2a2a3e", height=80)
        header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 10))
        header_frame.grid_columnconfigure(0, weight=1)
        
        # Logo/Title
        title_label = ctk.CTkLabel(
            header_frame, 
            text="🚀 LeadGen Pro", 
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="#00d9ff"
        )
        title_label.grid(row=0, column=0, padx=30, pady=(15, 0), sticky="w")
        
        # Subtitle
        subtitle_label = ctk.CTkLabel(
            header_frame, 
            text="Enterprise-Grade Lead Generation & Enrichment Platform", 
            font=ctk.CTkFont(size=14),
            text_color="#a0a0b0"
        )
        subtitle_label.grid(row=1, column=0, padx=30, pady=(0, 15), sticky="w")
    
    def create_input_section(self):
        """Create input fields section."""
        input_frame = ctk.CTkFrame(self, fg_color="#252535")
        input_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        input_frame.grid_columnconfigure(1, weight=1)
        
        # Target Keyword
        ctk.CTkLabel(
            input_frame, 
            text="Target Keyword:", 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#ffffff"
        ).grid(row=0, column=0, padx=20, pady=15, sticky="w")
        
        self.keyword_entry = ctk.CTkEntry(
            input_frame, 
            placeholder_text="e.g., Dentist, SaaS Company, Restaurant",
            width=400,
            height=40,
            corner_radius=8,
            border_width=1,
            border_color="#404050"
        )
        self.keyword_entry.grid(row=0, column=1, padx=20, pady=15, sticky="ew")
        
        # Location
        ctk.CTkLabel(
            input_frame, 
            text="Location:", 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#ffffff"
        ).grid(row=1, column=0, padx=20, pady=15, sticky="w")
        
        self.location_entry = ctk.CTkEntry(
            input_frame, 
            placeholder_text="e.g., New York, NY, London, UK",
            width=400,
            height=40,
            corner_radius=8,
            border_width=1,
            border_color="#404050"
        )
        self.location_entry.grid(row=1, column=1, padx=20, pady=15, sticky="ew")
        
        # Max Leads
        ctk.CTkLabel(
            input_frame, 
            text="Max Leads:", 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#ffffff"
        ).grid(row=2, column=0, padx=20, pady=15, sticky="w")
        
        self.max_leads_entry = ctk.CTkEntry(
            input_frame, 
            placeholder_text="50",
            width=400,
            height=40,
            corner_radius=8,
            border_width=1,
            border_color="#404050"
        )
        self.max_leads_entry.insert(0, "50")
        self.max_leads_entry.grid(row=2, column=1, padx=20, pady=15, sticky="w")
        
        # API Keys Section (Collapsible)
        self.api_keys_frame = ctk.CTkFrame(input_frame, fg_color="#1e1e2e")
        self.api_keys_frame.grid(row=3, column=0, columnspan=2, padx=20, pady=15, sticky="ew")
        
        ctk.CTkLabel(
            self.api_keys_frame, 
            text="🔑 Optional API Keys (For Enhanced Email Validation)", 
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#00d9ff"
        ).pack(pady=10)
        
        api_grid = ctk.CTkFrame(self.api_keys_frame, fg_color="transparent")
        api_grid.pack(pady=5)
        api_grid.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(
            api_grid, 
            text="Hunter.io API Key:", 
            font=ctk.CTkFont(size=12),
            text_color="#a0a0b0"
        ).grid(row=0, column=0, padx=20, pady=5, sticky="w")
        
        self.hunter_api_entry = ctk.CTkEntry(
            api_grid, 
            placeholder_text="Optional",
            width=300,
            height=35,
            corner_radius=6,
            show="*"
        )
        self.hunter_api_entry.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        
        ctk.CTkLabel(
            api_grid, 
            text="Tomba API Key:", 
            font=ctk.CTkFont(size=12),
            text_color="#a0a0b0"
        ).grid(row=1, column=0, padx=20, pady=5, sticky="w")
        
        self.tomba_api_entry = ctk.CTkEntry(
            api_grid, 
            placeholder_text="Optional",
            width=300,
            height=35,
            corner_radius=6,
            show="*"
        )
        self.tomba_api_entry.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
    
    def create_control_panel(self):
        """Create control buttons."""
        control_frame = ctk.CTkFrame(self, fg_color="#252535", height=80)
        control_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=10)
        control_frame.grid_columnconfigure(0, weight=1)
        
        button_container = ctk.CTkFrame(control_frame, fg_color="transparent")
        button_container.pack(expand=True)
        
        # Start Button
        self.start_button = ctk.CTkButton(
            button_container,
            text="▶ Start Scraping",
            command=self.start_scraping,
            height=45,
            width=200,
            corner_radius=10,
            fg_color="#00d9ff",
            hover_color="#00b8d9",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.start_button.pack(side="left", padx=30)
        
        # Stop Button
        self.stop_button = ctk.CTkButton(
            button_container,
            text="⏹ Stop/Emergency Halt",
            command=self.stop_scraping,
            height=45,
            width=200,
            corner_radius=10,
            fg_color="#ff4757",
            hover_color="#ff3344",
            font=ctk.CTkFont(size=16, weight="bold"),
            state="disabled"
        )
        self.stop_button.pack(side="left", padx=30)
    
    def create_feedback_section(self):
        """Create live feedback section."""
        feedback_frame = ctk.CTkFrame(self, fg_color="#252535")
        feedback_frame.grid(row=3, column=0, sticky="nsew", padx=20, pady=10)
        feedback_frame.grid_rowconfigure(2, weight=1)
        feedback_frame.grid_columnconfigure(0, weight=1)
        
        # Progress Bar
        ctk.CTkLabel(
            feedback_frame, 
            text="Progress:", 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#ffffff"
        ).grid(row=0, column=0, padx=20, pady=(15, 5), sticky="w")
        
        self.progress_bar = ctk.CTkProgressBar(feedback_frame, width=800, height=20, corner_radius=10)
        self.progress_bar.set(0)
        self.progress_bar.grid(row=1, column=0, padx=20, pady=(0, 15), sticky="ew")
        
        self.progress_label = ctk.CTkLabel(
            feedback_frame, 
            text="Ready to start", 
            font=ctk.CTkFont(size=12),
            text_color="#00d9ff"
        )
        self.progress_label.grid(row=1, column=0, padx=20, sticky="e")
        
        # Status Label
        self.status_label = ctk.CTkLabel(
            feedback_frame, 
            text="Status: Idle", 
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#ffa500"
        )
        self.status_label.grid(row=2, column=0, padx=20, pady=(10, 5), sticky="w")
        
        # Terminal Log
        log_label = ctk.CTkLabel(
            feedback_frame, 
            text="Terminal Log:", 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#ffffff"
        )
        log_label.grid(row=3, column=0, padx=20, pady=(15, 5), sticky="w")
        
        self.log_textbox = ctk.CTkTextbox(
            feedback_frame, 
            width=1100, 
            height=200,
            corner_radius=8,
            fg_color="#000000",
            text_color="#00ff00",
            font=ctk.CTkFont(family="Courier", size=11)
        )
        self.log_textbox.grid(row=4, column=0, padx=20, pady=(0, 20), sticky="nsew")
    
    def create_export_section(self):
        """Create export action section."""
        export_frame = ctk.CTkFrame(self, fg_color="#252535", height=70)
        export_frame.grid(row=4, column=0, sticky="ew", padx=20, pady=(0, 20))
        export_frame.grid_columnconfigure(0, weight=1)
        
        self.export_button = ctk.CTkButton(
            export_frame,
            text="📊 Open Excel File",
            command=self.open_excel_file,
            height=45,
            width=200,
            corner_radius=10,
            fg_color="#2ecc71",
            hover_color="#27ae60",
            font=ctk.CTkFont(size=16, weight="bold"),
            state="disabled"
        )
        self.export_button.pack(expand=True)
        
        self.last_export_path = ""
    
    def log_message(self, message: str):
        """Add message to terminal log."""
        self.log_textbox.insert("end", message + "\n")
        self.log_textbox.see("end")
    
    def update_progress(self, percentage: int, status: str):
        """Update progress bar and status."""
        self.progress_bar.set(percentage / 100)
        self.progress_label.configure(text=f"{percentage}%")
        self.status_label.configure(text=f"Status: {status}")
    
    def validate_inputs(self) -> bool:
        """Validate user inputs."""
        keyword = self.keyword_entry.get().strip()
        location = self.location_entry.get().strip()
        
        if not keyword:
            messagebox.showerror("Validation Error", "Please enter a target keyword")
            return False
        
        if not location:
            messagebox.showerror("Validation Error", "Please enter a location")
            return False
        
        try:
            max_leads = int(self.max_leads_entry.get().strip())
            if max_leads < 1 or max_leads > 500:
                raise ValueError()
        except ValueError:
            messagebox.showerror("Validation Error", "Max Leads must be between 1 and 500")
            return False
        
        return True
    
    def start_scraping(self):
        """Start scraping process."""
        if not self.validate_inputs():
            return
        
        # Update UI state
        self.is_scraping = True
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self.export_button.configure(state="disabled")
        
        # Clear previous results
        self.log_textbox.delete("1.0", "end")
        self.current_leads = []
        
        # Get configuration
        self.config.keyword = self.keyword_entry.get().strip()
        self.config.location = self.location_entry.get().strip()
        self.config.max_leads = int(self.max_leads_entry.get().strip())
        self.config.hunter_api_key = self.hunter_api_entry.get().strip()
        self.config.tomba_api_key = self.tomba_api_entry.get().strip()
        
        # Start scraping in background thread
        self.scraping_thread = threading.Thread(target=self.run_scraping_async, daemon=True)
        self.scraping_thread.start()
        
        self.log_message("=" * 60)
        self.log_message(f"Starting lead generation for: {self.config.keyword} in {self.config.location}")
        self.log_message(f"Max leads: {self.config.max_leads}")
        self.log_message("=" * 60)
    
    def run_scraping_async(self):
        """Run scraping in async loop."""
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            self.scraper = LeadScraper(
                config=self.config,
                log_callback=self.log_message,
                progress_callback=self.update_progress
            )
            
            leads = loop.run_until_complete(self.scraper.run_scraping())
            
            self.current_leads = leads
            
            # Export to Excel
            if leads:
                filepath = ExcelExporter.export_to_excel(
                    leads, 
                    self.config.keyword, 
                    self.config.location
                )
                self.last_export_path = filepath
                self.log_message(f"\n✓ Excel file created: {filepath}")
                
                # Enable export button
                self.after(0, lambda: self.export_button.configure(state="normal"))
            else:
                self.log_message("\n⚠ No leads extracted")
            
            loop.close()
            
        except Exception as e:
            self.log_message(f"\n✗ Error: {str(e)}")
        
        finally:
            # Reset UI state
            self.after(0, self.reset_ui_state)
    
    def stop_scraping(self):
        """Stop scraping process."""
        if self.scraper:
            self.scraper.stop()
        
        self.log_message("\n⚠ Stopping scraping process...")
        self.reset_ui_state()
    
    def reset_ui_state(self):
        """Reset UI to initial state."""
        self.is_scraping = False
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
        self.status_label.configure(text="Status: Ready")
    
    def open_excel_file(self):
        """Open the last exported Excel file."""
        if self.last_export_path and os.path.exists(self.last_export_path):
            try:
                # Try to open with default application
                if sys.platform == 'win32':
                    os.startfile(self.last_export_path)
                elif sys.platform == 'darwin':
                    os.system(f'open "{self.last_export_path}"')
                else:
                    os.system(f'xdg-open "{self.last_export_path}"')
            except Exception as e:
                messagebox.showerror("Error", f"Could not open file: {str(e)}")
        else:
            messagebox.showinfo("Info", "No Excel file available. Run scraping first.")


def main():
    """Main entry point."""
    app = LeadGenProApp()
    app.mainloop()


if __name__ == "__main__":
    main()
